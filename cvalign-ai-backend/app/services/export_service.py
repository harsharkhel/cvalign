import json
from io import BytesIO
from typing import Any, List

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models.ai_chat_message import AIChatMessage
from app.models.job import Job
from app.models.job_recommendation import JobRecommendation
from app.models.job_search_log import JobSearchLog
from app.models.login_log import LoginLog
from app.models.resume_analysis import ResumeAnalysis
from app.models.user import User

FORBIDDEN_EXPORT_COLUMNS = {
    "password",
    "password_hash",
    "access_token",
    "jwt",
    "google_token",
    "openai_api_key",
    "gemini_api_key",
    "secret",
}


def _workbook_to_bytes(wb: Workbook) -> BytesIO:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _add_sheet(wb: Workbook, title: str, headers: List[str], rows: List[List[Any]]):
    for col in headers:
        if col.lower() in FORBIDDEN_EXPORT_COLUMNS:
            raise ValueError(f"Forbidden export column: {col}")
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)


def export_users(db: Session) -> BytesIO:
    users = db.query(User).all()
    headers = [
        "user_uuid",
        "name",
        "email",
        "auth_provider",
        "role",
        "is_email_verified",
        "is_active",
        "created_at",
        "last_login",
    ]
    rows = [
        [
            u.user_uuid,
            u.name,
            u.email,
            u.auth_provider.value,
            u.role.value,
            u.is_email_verified,
            u.is_active,
            u.created_at,
            u.last_login,
        ]
        for u in users
    ]
    wb = Workbook()
    _add_sheet(wb, "users", headers, rows)
    return _workbook_to_bytes(wb)


def export_login_logs(db: Session) -> BytesIO:
    logs = db.query(LoginLog).order_by(LoginLog.login_time.desc()).all()
    headers = ["id", "user_id", "email", "provider", "status", "ip_address", "device_info", "login_time"]
    rows = [
        [
            l.id,
            l.user_id,
            l.email,
            l.provider.value,
            l.status.value,
            l.ip_address,
            l.device_info,
            l.login_time,
        ]
        for l in logs
    ]
    wb = Workbook()
    _add_sheet(wb, "login_logs", headers, rows)
    return _workbook_to_bytes(wb)


def export_resume_analyses(db: Session) -> BytesIO:
    items = db.query(ResumeAnalysis).all()
    headers = [
        "id",
        "user_id",
        "resume_filename",
        "job_title",
        "company_name",
        "ats_score",
        "text_similarity_score",
        "skill_match_score",
        "matched_skills",
        "missing_skills",
        "created_at",
    ]
    rows = [
        [
            a.id,
            a.user_id,
            a.resume_filename,
            a.job_title,
            a.company_name,
            a.ats_score,
            a.text_similarity_score,
            a.skill_match_score,
            json.dumps(a.matched_skills_json),
            json.dumps(a.missing_skills_json),
            a.created_at,
        ]
        for a in items
    ]
    wb = Workbook()
    _add_sheet(wb, "resume_analyses", headers, rows)
    return _workbook_to_bytes(wb)


def export_jobs(db: Session) -> BytesIO:
    jobs = db.query(Job).all()
    headers = [
        "id",
        "source",
        "title",
        "company",
        "location",
        "job_type",
        "remote_type",
        "apply_url",
        "source_url",
        "is_active",
        "created_at",
    ]
    rows = [
        [
            j.id,
            j.source,
            j.title,
            j.company,
            j.location,
            j.job_type.value,
            j.remote_type.value,
            j.apply_url,
            j.source_url,
            j.is_active,
            j.created_at,
        ]
        for j in jobs
    ]
    wb = Workbook()
    _add_sheet(wb, "jobs", headers, rows)
    return _workbook_to_bytes(wb)


def export_job_search_logs(db: Session) -> BytesIO:
    logs = db.query(JobSearchLog).all()
    headers = ["id", "user_id", "query", "location", "job_type", "source_used", "results_count", "searched_at"]
    rows = [
        [
            l.id,
            l.user_id,
            l.query,
            l.location,
            l.job_type,
            l.source_used,
            l.results_count,
            l.searched_at,
        ]
        for l in logs
    ]
    wb = Workbook()
    _add_sheet(wb, "job_search_logs", headers, rows)
    return _workbook_to_bytes(wb)


def export_job_recommendations(db: Session) -> BytesIO:
    recs = db.query(JobRecommendation).all()
    headers = [
        "id",
        "user_id",
        "resume_analysis_id",
        "job_id",
        "match_score",
        "matched_skills",
        "missing_skills",
        "recommendation_reason",
        "created_at",
    ]
    rows = [
        [
            r.id,
            r.user_id,
            r.resume_analysis_id,
            r.job_id,
            r.match_score,
            json.dumps(r.matched_skills_json),
            json.dumps(r.missing_skills_json),
            r.recommendation_reason,
            r.created_at,
        ]
        for r in recs
    ]
    wb = Workbook()
    _add_sheet(wb, "job_recommendations", headers, rows)
    return _workbook_to_bytes(wb)


def export_chat_messages(db: Session) -> BytesIO:
    messages = db.query(AIChatMessage).all()
    headers = ["id", "user_id", "resume_analysis_id", "job_id", "user_message", "ai_response", "created_at"]
    rows = [
        [
            m.id,
            m.user_id,
            m.resume_analysis_id,
            m.job_id,
            m.user_message,
            m.ai_response,
            m.created_at,
        ]
        for m in messages
    ]
    wb = Workbook()
    _add_sheet(wb, "ai_chat_messages", headers, rows)
    return _workbook_to_bytes(wb)
